#!/usr/bin/env python3
"""
Payslip Parser
Reads payslip PDFs from ./payslips/ and writes payslip_data.xlsx.

Usage:
    python3 parse.py                        # parse all PDFs
    python3 parse.py --debug FILE.pdf       # dump raw text from a PDF
"""

import re
import sys
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PAYSLIPS_DIR = Path(__file__).parent / "payslips"
OUTPUT_FILE = Path(__file__).parent / "payslip_data.xlsx"

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_amount(s):
    """'£1,234.56' or '-£1,234.56'  →  float, or None."""
    if s is None:
        return None
    s = str(s).strip().replace(',', '').replace('£', '').replace(' ', '')
    try:
        return float(s)
    except ValueError:
        return None


def get_text(pdf_path, page_num=0):
    with pdfplumber.open(pdf_path) as pdf:
        if page_num >= len(pdf.pages):
            return ''
        return pdf.pages[page_num].extract_text() or ''


def classify(name):
    lo = name.lower()
    if re.match(r'\d{4}-\d{2}-cap\.pdf$', lo):
        return 'payslip'
    if re.match(r'p60-\d{4}-\d{4}\.pdf$', lo):
        return 'p60'
    if any(x in lo for x in ('pay letter', 'reward letter', 'benefit funding')):
        return 'pay_letter'
    return 'other'


# ── Payslip ───────────────────────────────────────────────────────────────────

# Column x-boundaries (consistent across the supported payslip PDF layout)
_COL2_X = 200   # Deductions column starts here
_COL3_X = 370   # Totals & Balances column starts here
_Y_TOL  = 3     # pixels tolerance for grouping words into the same row

_CELL_AMOUNT = re.compile(r'^(.+?)\s+(-?£[\d,]+\.?\d*)$')


def _extract_columns(page):
    """
    Return list of (col1_text, col2_text, col3_text) tuples,
    one per visual row, using word x-coordinates to assign columns.
    """
    words = page.extract_words()

    # Group words into rows by y-coordinate
    rows: dict[float, list] = {}
    for w in words:
        y = round(w['top'] / _Y_TOL) * _Y_TOL
        rows.setdefault(y, []).append(w)

    result = []
    for y in sorted(rows):
        row_words = sorted(rows[y], key=lambda w: w['x0'])
        cols = ['', '', '']
        for w in row_words:
            x = w['x0']
            idx = 0 if x < _COL2_X else (1 if x < _COL3_X else 2)
            cols[idx] = (cols[idx] + ' ' + w['text']).strip()
        result.append(tuple(cols))
    return result


def _parse_cell(text):
    """'PENSION -£268.60' → ('PENSION', -268.60) or None."""
    if not text:
        return None
    m = _CELL_AMOUNT.match(text.strip())
    if m:
        return m.group(1).strip(), parse_amount(m.group(2))
    return None


def parse_payslip(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ''
        rows = _extract_columns(page)

    data = {'file': pdf_path.name, 'pay_items': {}}

    # Extract header fields from raw text (single-column "My details" section)
    for line in text.splitlines():
        m = re.search(r'Payslip Date\s+(\d{2}/\d{2}/\d{4})', line)
        if m:
            data['date'] = datetime.strptime(m.group(1), '%d/%m/%Y')
            data['year_month'] = data['date'].strftime('%Y-%m')
        m = re.search(r'Tax Code\s+(\S+)', line)
        if m:
            data['tax_code'] = m.group(1)

    # Find the header row and parse the data table
    in_table = False
    for c1, c2, c3 in rows:
        # Detect table header row
        if 'Pay & Allowances' in c1 and 'Deductions' in c2:
            in_table = True
            continue
        if not in_table:
            continue
        if 'No messages' in c1 or 'No messages' in c2:
            break

        # Parse each column cell
        p1 = _parse_cell(c1)
        p2 = _parse_cell(c2)
        p3 = _parse_cell(c3)

        # Column 1: Pay & Allowances
        if p1:
            name, amount = p1
            if name == 'TOTAL':
                data['gross_total'] = amount
            else:
                data['pay_items'][name] = amount

        # Column 2: Deductions
        if p2:
            name, amount = p2
            if name == 'Income Tax':
                data['income_tax_period'] = amount
            elif name == 'N.I':
                data['ni_period'] = amount
            elif name == 'TOTAL':
                data['total_deductions'] = amount

        # Column 3: Totals & Balances
        if p3:
            name, amount = p3
            if name == 'Income Tax':
                data['ytd_income_tax'] = amount
            elif name == 'Taxable Gross':
                data['ytd_taxable_gross'] = amount
            elif name == 'N.I':
                data['ytd_ni'] = amount
            elif name == 'Net Pay':
                data['net_pay'] = amount

    return data


# ── P60 ───────────────────────────────────────────────────────────────────────

def parse_p60(pdf_path):
    text = get_text(pdf_path, 0)
    flat = ' '.join(text.split())
    data = {'file': pdf_path.name}

    m = re.match(r'p60-(\d{4}-\d{4})\.pdf', pdf_path.name, re.IGNORECASE)
    if m:
        data['tax_year'] = m.group(1)

    m = re.search(r'In this employment\s*\*?\s*£([\d,]+\.?\d*)\s+£([\d,]+\.?\d*)', flat)
    if m:
        data['pay'] = parse_amount(m.group(1))
        data['tax_deducted'] = parse_amount(m.group(2))

    m = re.search(r'Gross Pay\s+£([\d,]+\.?\d*)', flat)
    if m:
        data['gross_pay'] = parse_amount(m.group(1))

    # NI row: "A £6,396 £6,180 £37,692 £3,484.09"
    m = re.search(r'\bA\s+£([\d,]+)\s+£([\d,]+)\s+£([\d,]+)\s+£([\d,]+\.\d+)', flat)
    if m:
        data['ni_lel']           = parse_amount(m.group(1))
        data['ni_lel_pt']        = parse_amount(m.group(2))
        data['ni_pt_uel']        = parse_amount(m.group(3))
        data['ni_contributions'] = parse_amount(m.group(4))

    m = re.search(r'Final tax code\s+(\S+)', flat)
    if m:
        data['final_tax_code'] = m.group(1)

    return data


# ── Pay / Reward Letters ──────────────────────────────────────────────────────

def parse_pay_letter(pdf_path):
    text = get_text(pdf_path, 0)
    flat = ' '.join(text.split())
    lo   = pdf_path.name.lower()
    data = {'file': pdf_path.name}

    # Benefit funding change letter
    if 'benefit funding' in lo:
        data['type'] = 'Benefit Change'
        m = re.search(r'Cash Allowance will be £\s*([\d,]+)', flat)
        if m:
            data['new_cash_allowance'] = parse_amount(m.group(1))
        m = re.search(r'increased by £\s*([\d,]+\.?\d*)', flat)
        if m:
            data['benefit_increase_amount'] = parse_amount(m.group(1))
        # "effective from X" or "effective X"
        m = re.search(r'effective (?:from )?(\d+\w* \w+ \d{4})', flat)
        if m:
            data['effective_date_text'] = m.group(1)
        return data

    data['type'] = 'Q1 Pay Review' if 'q1' in lo else (
                   'Reward / Promotion' if 'reward' in lo else 'Pay Review')

    # Effective date
    m = re.search(r'from (?:the )?(\d+(?:st|nd|rd|th)? \w+ \d{4})', flat)
    if m:
        data['effective_date_text'] = m.group(1)

    # Table: Base Pay / Benefit Funding / Reward Package
    for label in ('Base Pay', 'Benefit Funding', 'Reward Package'):
        key = label.lower().replace(' ', '_')
        m = re.search(
            rf'{label}\s+£([\d,]+)\s+£([\d,]+)\s+£([\d,]+)\s+([\d.]+)%',
            flat
        )
        if m:
            data[f'{key}_current']  = parse_amount(m.group(1))
            data[f'{key}_new']      = parse_amount(m.group(2))
            data[f'{key}_increase'] = parse_amount(m.group(3))
            data[f'{key}_pct']      = float(m.group(4))

    return data


# ── Excel helpers ─────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_HDR_FONT = Font(color='FFFFFF', bold=True)
_ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
_GBP      = '£#,##0.00'


def _style_headers(ws, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 30


def _autowidth(ws, max_w=28):
    for col in ws.columns:
        w = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)


def _alt_row(ws, row, n_cols):
    if row % 2 == 0:
        for c in range(1, n_cols + 1):
            ws.cell(row=row, column=c).fill = _ALT_FILL


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_payslips(wb, payslips):
    ws = wb.create_sheet('Monthly Payslips')

    # Collect all pay item names across every payslip (preserving first-seen order)
    seen_items: dict[str, None] = {}
    for p in payslips:
        for k in p['pay_items']:
            seen_items.setdefault(k, None)
    dynamic = [k for k in seen_items if k != 'BASE PAY']

    fixed   = ['Year-Month', 'Date', 'Tax Code', 'Base Pay']
    tail    = ['Gross Total', 'Income Tax (Period)', 'NI (Period)',
               'Total Deductions', 'Net Pay',
               'YTD Taxable Gross', 'YTD Income Tax', 'YTD NI']
    headers = fixed + dynamic + tail

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_headers(ws, len(headers))

    for r, p in enumerate(payslips, 2):
        c = 1
        ws.cell(row=r, column=c, value=p.get('year_month')); c += 1

        cell = ws.cell(row=r, column=c, value=p.get('date'))
        if p.get('date'):
            cell.number_format = 'DD/MM/YYYY'
        c += 1

        ws.cell(row=r, column=c, value=p.get('tax_code')); c += 1
        ws.cell(row=r, column=c, value=p['pay_items'].get('BASE PAY')); c += 1

        for item in dynamic:
            ws.cell(row=r, column=c, value=p['pay_items'].get(item)); c += 1

        for key in ('gross_total', 'income_tax_period', 'ni_period',
                    'total_deductions', 'net_pay',
                    'ytd_taxable_gross', 'ytd_income_tax', 'ytd_ni'):
            ws.cell(row=r, column=c, value=p.get(key)); c += 1

        _alt_row(ws, r, len(headers))

    # Apply currency format to numeric cells from col 4 onwards
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = _GBP

    _autowidth(ws)
    ws.freeze_panes = 'D2'


def write_p60(wb, p60s):
    ws = wb.create_sheet('P60 Annual')
    headers = ['Tax Year', 'Pay in Employment', 'Tax Deducted', 'Gross Pay',
               'NI Earnings (LEL)', 'NI Earnings (LEL→PT)', 'NI Earnings (PT→UEL)',
               'Employee NI Contributions', 'Final Tax Code']

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_headers(ws, len(headers))

    for r, p in enumerate(p60s, 2):
        vals = [p.get('tax_year'), p.get('pay'), p.get('tax_deducted'), p.get('gross_pay'),
                p.get('ni_lel'), p.get('ni_lel_pt'), p.get('ni_pt_uel'),
                p.get('ni_contributions'), p.get('final_tax_code')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if isinstance(v, float):
                cell.number_format = _GBP
        _alt_row(ws, r, len(headers))

    _autowidth(ws)
    ws.freeze_panes = 'B2'


def write_letters(wb, letters):
    ws = wb.create_sheet('Pay & Reward Letters')
    headers = ['File', 'Type', 'Effective Date',
               'Base Pay (Current)', 'Base Pay (New)', 'Base Pay Increase', 'Base Pay %',
               'Benefit Funding (Current)', 'Benefit Funding (New)', 'Benefit Funding Increase',
               'Reward Package (Current)', 'Reward Package (New)', 'Reward Package Increase',
               'New Cash Allowance', 'Benefit Increase Amount']
    currency_cols = {4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15}

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_headers(ws, len(headers))

    for r, p in enumerate(letters, 2):
        row_vals = [
            p.get('file'), p.get('type'), p.get('effective_date_text'),
            p.get('base_pay_current'), p.get('base_pay_new'), p.get('base_pay_increase'),
            p.get('base_pay_pct'),
            p.get('benefit_funding_current'), p.get('benefit_funding_new'),
            p.get('benefit_funding_increase'),
            p.get('reward_package_current'), p.get('reward_package_new'),
            p.get('reward_package_increase'),
            p.get('new_cash_allowance'), p.get('benefit_increase_amount'),
        ]
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if isinstance(v, float) and c in currency_cols:
                cell.number_format = _GBP
        _alt_row(ws, r, len(headers))

    _autowidth(ws, max_w=40)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) >= 3 and sys.argv[1] == '--debug':
        path = PAYSLIPS_DIR / sys.argv[2] if not Path(sys.argv[2]).is_absolute() else Path(sys.argv[2])
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                print(f'=== Page {i} ===')
                print(page.extract_text())
        return

    payslips, p60s, letters = [], [], []
    pdfs = sorted(PAYSLIPS_DIR.glob('*.pdf'))
    print(f'Found {len(pdfs)} PDF files\n')

    for path in pdfs:
        doc_type = classify(path.name)
        try:
            if doc_type == 'payslip':
                d = parse_payslip(path)
                payslips.append(d)
                status = f"date={d.get('year_month')}  net_pay={d.get('net_pay')}"
            elif doc_type == 'p60':
                d = parse_p60(path)
                p60s.append(d)
                status = f"year={d.get('tax_year')}  pay={d.get('pay')}"
            elif doc_type == 'pay_letter':
                d = parse_pay_letter(path)
                letters.append(d)
                status = f"type={d.get('type')}  eff={d.get('effective_date_text')}"
            else:
                status = 'skipped'
            print(f'  [{doc_type:10s}] {path.name:45s} {status}')
        except Exception as e:
            print(f'  [ERROR     ] {path.name}: {e}')

    payslips.sort(key=lambda x: x.get('date') or datetime.min)
    p60s.sort(key=lambda x: x.get('tax_year', ''))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_payslips(wb, payslips)
    write_p60(wb, p60s)
    write_letters(wb, letters)
    wb.save(OUTPUT_FILE)

    print(f'\nSaved → {OUTPUT_FILE}')
    print(f'  Monthly payslips  : {len(payslips)}')
    print(f'  P60 certificates  : {len(p60s)}')
    print(f'  Pay/reward letters: {len(letters)}')


if __name__ == '__main__':
    main()
